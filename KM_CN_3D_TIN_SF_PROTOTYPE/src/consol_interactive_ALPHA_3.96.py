# Libraries 
import os
import glob
import pandas as pd
import numpy as np
import math
from pandas import DataFrame
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")
import win32com.client
from win32com import client
import xlsxwriter


# Section-0

# Project_Globa_Directory must be set up 

cathi = input("This is a Windows-Based script. Please enter the path to the project directory (e.g. C:/Users/..):   ")
prnamepv = input('Project Name:  ')
querytagA = int(input("Enter the 1st tag for DX: "))
querytagB = int(input("Enter the last tag for DX: "))
inp = cathi+"/Inputs/"

# Section-1

# The users enter the datasets names (e.g. drawings numbers, projects) and a range of Tag numbers that represent the road-sections
## If we proceed to the production phase, we will store all drawings data from a project in one mega-database 

pv = pd.DataFrame()
for file in glob.glob(os.path.join(inp,"plot*.csv")):
    df = pd.read_csv(file,index_col=0)
    if pv.empty:
        pv = df
    else:
        pv = pv.join(df, how='outer')

dfpv = pv

lv = pd.DataFrame()
for filee in glob.glob(os.path.join(inp,"longi*.csv")):
    dfd = pd.read_csv(filee,index_col=0)
    if lv.empty:
        lv = dfd
    else:
        lv = lv.join(dfd, how='outer')

dflv = lv

# Activate for Ubuntu 
#cathy = '/mnt/c/Users/CINPC0075/Desktop/Repo-KCIN3D-reconstruction/For_UI_FS'
# Activate for Windows 
#cathy = r'C:\Users\CINPC0075\Desktop\Repo-KCIN3D-reconstruction\For_UI_FS'cd'

cathy = cathi
mdpathpv_OUT0 = cathy+ '/'+ 'Outputs'+'/'+ prnamepv+"_OUT0" +'.xlsx'
mdpathpv_OUT1 = cathy+ '/'+ 'Outputs'+'/'+ prnamepv+"_OUT1" +'.xlsx'
mdpathpv_OUT2 = cathy+ '/'+ 'Outputs'+'/'+ prnamepv+"_OUT2" +'.xlsx'
mdpathpv_OUT10 = cathy+ '/'+'Outputs'+'/'+ prnamepv+"_OUT0M" +'.xlsx'
mdpathpv_OUT1M = cathy+ '/'+ 'Outputs'+'/'+ prnamepv+"_OUT1M" +'.xlsm'
mdpathpv_OUT2M = cathy+ '/'+ 'Outputs'+'/'+ prnamepv+"_OUT2M" +'.xlsm'
source = cathy+ '/'+ 'src'+'/'+'sbinCAD.xlsm'
generator = cathy+ '/'+ 'Temp'+'/'+ 'generator.xlsm'
generator1 = cathy+ '/'+ 'Temp'+'/'+ 'generatorCAD.xlsm'
generator2 = cathy+ '/'+ 'Temp'+'/'+ 'generatorCADIN.xlsm'


# Section-2

# Here we do some arrangements, parameter-setting and cleaning the header
to_drop = ['Author','Closed', 'Comments', 'Drawing Revision Number', 'EdgeStyleId', 'FaceStyleId', 'File Accessed', 'File Created', 'File Last Saved By', 'File Location',
           'File Modified','File Name', 'File Size', 'Fit/Smooth', 'Hyperlink', 'Hyperlink Base', 'Keywords']
dfpv.drop(to_drop, inplace=True, axis=1)
dflv.drop(to_drop, inplace=True, axis=1)

dfpv.set_index("Value",inplace= True)
dflv.set_index("Value",inplace= True)
rangi = querytagB - querytagA

# Section-3

# Extracting the coordinates x,y associated with Tag-Numbers (e.g. NO.270). The range has been specified by the users in Section-1, line-3 and 4
dtag = []
ran = range(querytagA,querytagB)
for i in ran:
    m = "NO."+str(i)
    dtag.append(m)
row = dtag

dff = dfpv.loc[row, ["Position X", "Position Y", "Position Z"]]
dff.sort_values("Position X", inplace=True, ascending=True)
## dff contains X,Y of each Tag-number ............... For details, see Work_ON_DX_Komatsu.ipynb
### the information in dff will help us find the coordinates of each point along the road-centerline 
#### Tag-numbers are major keys for us to connect different drawings together and extract relevant information


# Section-4

# Here, we do Euclidean distance similarity measure to find the pair (Tag-number, corresponding-road-centerline-point)
## We also extract the pair (Tag-Number, corresponding-circle-along-road-centerline)
df1 = dfpv.loc[dfpv["Name"] == 'Circle']
df2 = df1.loc[df1["Radius"] == 2.5]
df22 = df2.loc[df2["Name"] == "Circle", ["Center X", "Center Y", "Center Z"]]
df22.sort_values("Center X", inplace=True, ascending=True)
rrx = df22["Center X"]
rry = df22["Center Y"]
rrrx = dff["Position X"]
rrry = dff["Position Y"]
rrrr0 = dff.index
fel=[]
for i, j in zip(rrx,rry):
    for m, n, k in zip(rrrx,rrry,rrrr0):
        f= math.sqrt((j-n)**2 + (i-m)**2)
        fel.append([i,j,m,n,k,f])
d1 = pd.DataFrame(fel, columns=['CX','CY','TX','TY','TG','Dis'])
d11 = d1[:30]
d110 = d11.sort_values(by='Dis', ascending=True)
d11out=d110[:1]
d12 = d1[30:60]
d120 = d12.sort_values(by='Dis', ascending=True)
d12out= d120[:1]
d13 = d1[60:90]
d130 = d13.sort_values(by='Dis', ascending=True)
d13out = d130[:1]
d14 = d1[90:120]
d140 = d14.sort_values(by='Dis', ascending=True)
d14out = d140[:1]
d15 = d1[120:150]
d150 = d15.sort_values(by='Dis', ascending=True)
d15out = d150[:1]
d16 = d1[150:180]
d160 = d16.sort_values(by='Dis', ascending=True)
d16out = d160[:1]
d17 = d1[180:210]
d170 = d17.sort_values(by='Dis', ascending=True)
d17out = d170[:1]
dframe =  [d11out,d12out,d13out,d14out,d15out,d16out,d17out]
fi = pd.concat(dframe)
fi.reset_index(drop=True, inplace=True)
fi.sort_values("CX", inplace=True, ascending=True)
fi.insert(0, 'GEO', 'Circle')
fi.set_index('GEO',inplace= True)
fi["P1"] = fi["CX"].map(str)+','+fi["CY"].map(str)
fi.insert(7, 'P2', '1.5')
dl1 = dfpv.loc[dfpv["Name"] == 'Line']
dl2 = dl1.loc[dl1['Color'] == "red"]
dl3 = dl2.loc[dl2["Length"] == 5]
dl4 = dl3.loc[dl3['Name'] == "Line", ["End X", "End Y", "End Z",'Start X','Start Y','Start Z']]
dl4.reset_index(drop=True, inplace=True)
dl4.sort_values("End X", inplace=True, ascending=True)
dl4.insert(0, 'GEO', 'Line')
dl4.set_index('GEO',inplace= True)
dl4["P1"] = dl4["End X"].map(str)+','+dl4["End Y"].map(str)+','+dl4["End Z"].map(str)
dl4["P2"] = dl4["Start X"].map(str)+','+dl4["Start Y"].map(str)+','+dl4["Start Z"].map(str)
dl41 = dl4
dl5 = dl41
dl55= dl5[:rangi]
clinepointx = []
clinepointy = []
ran1 = dl55['End X'].astype(float)
ran2 = dl55['Start X'].astype(float)
ran3 = dl55['End Y'].astype(float)
ran4 = dl55['Start Y'].astype(float)
for ik,jk,kk,lk in zip(ran1,ran2,ran3,ran4):
    mnu = (ik+jk)/2
    mnv = (kk+lk)/2
    clinepointx.append(mnu)
    clinepointy.append(mnv)
dl55.insert(8, 'Cen X', clinepointx)
dl55.insert(9, 'Cen Y', clinepointy)
rown = pd.Series(row).str.replace('NO.', '', regex=False)
rowny = rown.values.tolist()
dl55.insert(10, 'Value', rowny)
dl55.drop(['P1','P2'], inplace=True, axis=1)
dl4l=dl9l = dl4[:rangi]
dl4l.insert(8, 'Value', row)
dll4 = dl4.drop(["End X", "End Y", "End Z",'Start X','Start Y','Start Z'], axis=1)
fii = fi.drop(["CX", "CY", "TX",'TY','TG','Dis'], axis=1)
wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(dll4, index=True, header=True):
    ws.append(r)
for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'
#wb.save(mdpathpv_OUT0)
wba = Workbook()
wsa = wba.active
for k in dataframe_to_rows(fii, index=True, header=True):
    wsa.append(k)
for cell in wsa['A'] + wsa[1]:
    cell.style = 'Pandas'
#wba.save(mdpathpv_OUT1)

##### The ultimate outputs of Sectio-4 are the coordinations (X,Y,Z) of all designated points along the road-centerline (Stored in a xlsx file for re-creating and 
#####   drawing the road-centerline by use of VBA)
##### RECALL, The true value of Z for the points must be extracted from Profile-view, so we will find the true elevation and replace Z with the true value 
##### RECALL, we use Tag-numbers to extract the true elevation and value of Z for each datapoint through the following sections 


# Section-5

# here, we use the Tag-numbers information and Profile-view drawings to extract the true values for Z

## these lines of code help us find the box in the drawings that contains true elevations
ref0 = dflv.loc[["高", "計","画"], ["Position X", "Position Y", "Position Z"]]
r1 = ref0['Position Y'].min()
r2 = ref0['Position Y'].max()
lm = dflv.loc[dflv["Position Y"]>(r1+8.491)]
lq = lm.loc[lm["Position Y"]<(r1+22.491)]
lz = lq.loc[:,["Position X","Position Y","Position Z"]]
dtemp1 = dflv.reset_index()
dtemp2 = dtemp1.dropna(subset=['Value'])

## we get the grand reference information
dtemp3 = dtemp2[dtemp2['Value'].str.contains('DL=')]
dtemp4 = dtemp3.iat[0,0]
reflev = float(dtemp4.replace('DL=',''))


## we extract the true elevations for each point
lz1 = lz.drop(["高", "計","画"])
lz2 = lz1.reset_index()
lz2.sort_values('Position X', inplace=True, ascending=True)
lz22 = lz2.loc[:,["Position X"]]
lz2['Value'] = lz2['Value'].astype(float)
lz3 = lz2['Value'] - reflev 
lz22.insert(0, 'TRUE ELEVATION', lz3)
b1 = lz22['Position X'].min()
b2 = lz22['Position X'].max()

## we extract and concatenate the true elevations with the corresponding Tag-numbers
dtemp33 = dtemp2[dtemp2['Value'].str.contains('NO.')]
dtemp333 = dtemp33.loc[dtemp33['Position X']>= b1]
dtemp444 = dtemp333.loc[dtemp333['Position X']<= b2]
dtemp444.sort_values('Position X', inplace=True, ascending=True)
dTG0 = dtemp444.loc[:,['Value','Position X']]
dTG = dTG0.loc[dTG0["Value"].apply(len)<7]
dTG['Value'] = dTG.Value.str.replace(r'(NO.)', '')
z1 = lz22.set_index("Position X")
z2 = dTG.set_index("Position X")
dTGLEV=pd.merge(z1, z2, left_index=True, right_index=True, how='outer')
tt = dTGLEV.dropna()
dimi = tt.loc[tt['Value']>=str(querytagA)]
dimit = dimi.loc[dimi['Value']<str(querytagB)]
dimitri = dimit.rename(columns={"TRUE ELEVATION": "End Z", 'Value': "Value"})
dl5501 = dl55.loc[dl55['Value']<str(querytagB)]
dimitri.reset_index()
dimitri.set_index('Value')
dl5501.reset_index()
dl5501.set_index('Value')
dfin = (dl5501.merge(dimitri, on='Value', how='left'))
dfin['Start Z'] = dfin["End Z_y"].add(dfin['Start Z'], fill_value=0)
dfinal = dfin.drop(['End Z_y', 'End Z_x'],axis=1)
# Intermed. temp file for calculation
ganzo = cathy+'/'+'Temp/'+'dfinal.pkl' 
dfinal.to_pickle(ganzo)

# LINE - PLANE-ST
dLINE = pd.read_pickle(ganzo)
dLINE.insert(0, 'GEO', 'Line')
dLINE.set_index('GEO',inplace= True)
dLINE["P1"] = dLINE["End X"].map(str)+','+dLINE["End Y"].map(str)+','+dLINE["Start Z"].map(str)
dLINE["P2"] = dLINE["Start X"].map(str)+','+dLINE["Start Y"].map(str)+','+dLINE["Start Z"].map(str)
dLINE.drop(['End X', 'End Y', 'Start X','Start Y', 'Start Z','Cen X', 'Cen Y', "Value"], inplace=True, axis=1)
wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(dLINE, index=True, header=True):
    ws.append(r)
for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'
#wb.save(mdpathpv_OUT0)
wbb = openpyxl.load_workbook(source, keep_vba=True)
sheet = wbb['Sheet1']
for r in dataframe_to_rows(dLINE, index=True, header=True):
    sheet.append(r)
for cell in sheet['A'] + sheet[1]:
    cell.style = 'Pandas'
wbb.save(generator)


# Road CentertLINE - Poly
dPLINE = pd.read_pickle(ganzo)
dPLINE["3DPOLY"] = dPLINE["Cen X"].map(str)+','+dPLINE["Cen Y"].map(str)+','+dPLINE["Start Z"].map(str)
dPLINE.drop(['End X', 'End Y', 'Start X','Start Y', 'Start Z','Cen X', 'Cen Y', "Value"], inplace=True, axis=1)
dTEX = dPLINE.T
wbp = Workbook()
wsp = wbp.active
for rt in dataframe_to_rows(dTEX, index=True, header=True):
    wsp.append(rt)
for cell in wsp['A'] + wsp[1]:
    cell.style = 'Pandas'
#wbp.save(mdpathpv_OUT2)


wbaa = openpyxl.load_workbook(generator, keep_vba=True)
sheet0 = wbaa['Sheet1']
for r in dataframe_to_rows(dTEX, index=True, header=True):
    sheet0.append(r)
for cell in sheet0['A'] + sheet0[1]:
    cell.style = 'Pandas'
wbaa.save(generator1)


EXCEL_CLS_NAME = "Excel.Application"

class XlMacro:
    def __init__(self, path, book, module, name, *args):
        self._path = path  # path containing workbook
        self._book = book  # workbook name like Book1.xlsm
        self._module = module  # module name, e.g., Module1
        self._name = name  # procedure or function name
        self._params = args  # argument(s)
        self._wb = None
    @property
    def workbook(self):
        return self._wb
    @property
    def wb_path(self):
        return os.path.join(self._path, self._book)
    @property
    def name(self):
        return f'{self._book}!{self._module}.{self._name}'
    @property
    def params(self):
        return self._params
    def get_workbook(self):
        wb_name = os.path.basename(self.wb_path)
        try:
            xl = client.GetActiveObject(EXCEL_CLS_NAME)
        except:
            # Excel is not running, so we need to handle it.
            xl = client.Dispatch(EXCEL_CLS_NAME)
        if wb_name in [wb.Name for wb in xl.Workbooks]:
            return xl.Workbooks[wb_name]
        else:
            return xl.Workbooks.Open(self.wb_path)
    def Run(self, *args, **kwargs):
        """ 
        Runs an Excel Macro or evaluates a UDF 
        """
        keep_open = kwargs.get('keep_open', True)
        save_changes = kwargs.get('save_changes', False)
        self._wb = self.get_workbook()
        xl_app = self._wb.Application
        xl_app.Visible = True
        ret = None
        if args is None:
            ret = xl_app.Run(self.name)
        elif not args:
            # run with some default parameters
            ret = xl_app.Run(self.name, *self.params)
        else:
            ret = xl_app.Run(self.name, *args)
        if not keep_open:
            self.workbook.Close(save_changes)
            self._wb = None
            xl_app.Quit()
        return ret

# Grand File-Output-1
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
starto = rangi+3
#append_df_to_excel(mdpathpv_OUT0, dTEX , sheet_name='Sheet', startrow=starto, truncate_sheet=False)

path = cathy+'/'+'Temp'+'/'
book2 = '183D1PLZ_平面図04_OUT0.xlsm'
book1 = '183D1PLZ_平面図04_OUT2M.xlsm'
book = 'generatorCAD.xlsm'
module = 'Module1'
macros = ['GetmyDatatoAutoCADC']
def default_params(macro):
    d = {'GetmyDatatoAutoCADC': []
        }
    return d.get(macro)



# Hamilton Quaternion Translation 3D

def divide_chunks(l, n):
    for i in range(0, len(l), n):
        yield l[i:i + n]

def PointRotate3D(p1, p2, p0, theta):
    from math import cos, sin, sqrt

    # Translate so axis is at origin    
    p = np.subtract(p0,p1)
    # Initialize point q
    q = (0.0,0.0,0.0)
    N = np.subtract(p2,p1)
    Nm = math.sqrt(N[0]**2+N[1]**2+N[2]**2)
    
    # Rotation axis unit vector
    n = (N[0]/Nm, N[1]/Nm, N[2]/Nm)

    # Matrix common factors     
    c = cos(theta)
    t = (1 - cos(theta))
    s = sin(theta)
    X = n[0]
    Y = n[1]
    Z = n[2]

    # Matrix 'M'
    d11 = t*X**2 + c
    d12 = t*X*Y - s*Z
    d13 = t*X*Z + s*Y
    d21 = t*X*Y + s*Z
    d22 = t*Y**2 + c
    d23 = t*Y*Z - s*X
    d31 = t*X*Z - s*Y
    d32 = t*Y*Z + s*X
    d33 = t*Z**2 + c

    #            |p.x|
    # Matrix 'M'*|p.y|
    #            |p.z|
    q0 = d11*p[0]+ d12*p[1]+ d13*p[2]
    q1 = d21*p[0]+ d22*p[1]+ d23*p[2]
    q2 = d31*p[0]+ d32*p[1]+ d33*p[2]
    
    q = (q0,q1,q2)

    # Translate axis and rotated point back to original location    
    return np.sum([[q],[p1]],axis=0)

def GenerateAssembly(p1, p2, p3, ops):
    zami = []
    for x in ops:
        zi = list(divide_chunks(x, 3))
        for y in zi:
            p0 = y 
            zii = PointRotate3D(p1, p2, p0, theta).tolist()
            for z in zii:
                p00 = z
                ziii = PointRotate3D(p1, p3, p00, 90)
                ziiio = ziii.tolist()
                zami.append(ziiio)
    a = np.array(zami)
    mat=np.matrix(a)
    dtra = DataFrame(mat)
    dtra
    dtraO=dtra[dtra.index%2==0]
    dtraO.reset_index(drop=True, inplace=True)
    dtraE=dtra[dtra.index%2==1]
    dtraE.reset_index(drop=True, inplace=True)
    dtraRE = pd.concat([dtraO, dtraE], axis=1)
    return dtraRE

# Translation 3D Coplanar det.

inp = cathi + '/Inputs/'
dassem = pd.read_pickle(cathi + "/Temp/dfinal.pkl")
dassemb = pd.read_pickle(cathi+ "/Temp/dfinal.pkl")

clinepointeta = []
rana1 = dassem['End X'].astype(float)
rana2 = dassem['Cen X'].astype(float)
rana3 = dassem['End Y'].astype(float)
rana4 = dassem['Cen Y'].astype(float)
for iko,jko,kko,lko in zip(rana1,rana2,rana3,rana4):
    mnu = (iko-jko)
    mmu = (kko-lko)
    mnv = math.sqrt(mmu**2+mnu**2)
    teta = np.arccos(mnu/mnv)
    clinepointeta.append(teta)
dassem.insert(8, 'Teta', clinepointeta)
dassem.drop(['End X', 'End Y', 'Start X','Start Y'], inplace=True, axis=1)
dassem = dassem[['Cen X', 'Cen Y','Start Z','Teta']]
dassemb.drop(['Cen X', 'Cen Y', 'Start X','Start Y','Value'], inplace=True, axis=1)
dassemb = dassemb[['End X', 'End Y','Start Z']]
######
######
######
dani = len(dassem)-1
rangROVER = range(querytagA,querytagB,1)
landROVER = range(0,dani,1)
for i , il in zip(rangROVER,landROVER):
    fileee = inp+"assem_"+ str(i) + ".csv"
    dfr = pd.read_csv(fileee, error_bad_lines=False, warn_bad_lines=False)
    to_drop = ['Author','Closed', 'Comments', 'Drawing Revision Number', 'EdgeStyleId', 'FaceStyleId', 'File Accessed', 'File Created', 'File Last Saved By', 'File Location',
           'File Modified','File Name', 'File Size', 'Fit/Smooth', 'Hyperlink', 'Hyperlink Base', 'Keywords']
    dfr.drop(to_drop, inplace=True, axis=1)
    ar = dfr.loc[dfr["Name"] == "Arc"]
    ar.loc[:,["Center X"]]
    lmn = dfr.loc[dfr["Name"]== "Line"]
    llm1 = lmn.loc[lmn["End X"]==lmn["Start X"]]
    mpa =0
    for xi in ar["Center X"]:
        for xii in llm1["End X"]:
            if xi == xii:
                mpa=xii

    dtemp10 = dfr.reset_index() 
    dtemp11 = dtemp10.dropna(subset=['Value'])
    dtemp33 = dtemp11[dtemp11['Value'].str.contains('DL=')]
    temp22 = dtemp33.iloc[0]['Position Y']
    temp33 = dtemp33.iloc[0]['Value']
    temp44 = float(temp33.replace('DL=',''))
    lq = lmn.loc[lmn["Color"]=="red"]
    lqz = lq.loc[:, ["End X", "End Y", "End Z", "Start X", "Start Y", "Start Z"]]
    lqx = lqz.sort_values(by= 'Start X', ascending=True)
    lqtest = lqx.loc[:,["End X", "End Y", "End Z", "Start X", "Start Y", "Start Z"]]
    #for 3D polyline half the road right
    
    lqhalfR = lqtest.loc[lqtest["Start X"]>= mpa]
    lqhalfR.loc[:, ["End X", "End Y", "End Z", "Start X", "Start Y", "Start Z"]]
    llq00 = lqhalfR.loc[:,["Start X", "Start Y", "Start Z"]]
    llqR = llq00.sort_values(by= 'Start X', ascending=True)
    
    lqhalfL= lqtest.loc[lqtest["Start X"]< mpa]
    lqhalfL.loc[:, ["End X", "End Y", "End Z", "Start X", "Start Y", "Start Z"]]
    llq0 = lqhalfL.loc[:,["Start X", "Start Y", "Start Z"]]
    llqL = llq0.sort_values(by= 'Start X', ascending=False)
                                          
    llqRR = lqhalfR.loc[:, ["End X", "End Y", "End Z", "Start X", "Start Y", "Start Z"]]
    f_row = llqR.iloc[[0]].values[0]
    translate = np.concatenate([f_row, f_row])
    llq_0 = llqRR.apply(lambda row: row - translate, axis=1)
    dassem_0 = dassem.loc[[il]].astype(float)                                          
    
    dassem_1 = dassem_0.iloc[[0]].values[0]
    index = 3
    ao = np.delete(dassem_1, index)
    translate2 = np.concatenate([ao,ao])
    llq_1R = llq_0.apply(lambda row: row + translate2, axis=1)

    llqLL = lqhalfL.loc[:, ["End X", "End Y", "End Z", "Start X", "Start Y", "Start Z"]]
    f_rowL = llqL.iloc[[0]].values[0]
    translateL = np.concatenate([f_rowL, f_rowL])
    llqLL_1 = llqLL.sort_values(by= 'Start X', ascending=False)
    llq_0L = llqLL_1.apply(lambda row: row - translateL, axis=1)
    
    dassem_1L = dassem_0.iloc[[0]].values[0]
    index = 3
    aoL = np.delete(dassem_1L, index)
    translate2L = np.concatenate([aoL,aoL])
    llq_1L = llq_0L.apply(lambda row: row + translate2L, axis=1)

    uu_0 = dassem.loc[[il]].astype(float)
    v = dassemb.loc[[il]].astype(float)
    theta = uu_0['Teta']
    u = uu_0[['Cen X', 'Cen Y','Start Z']]
    uu=u.iloc[0].values
    vv = v.iloc[0].values
    
    ww_0 = dassem.loc[[il]].astype(float)
    w = ww_0[['Cen X', 'Cen Y','Start Z']]
    if  ww_0.iloc[0]["Start Z"]>0:
        w['Start Z'] = w['Start Z'] - w['Start Z']
    else:
        w['Start Z'] = w['Start Z'] - 1
    ww = w.iloc[0].values
    
    p1= uu
    p2= ww
    p3= vv
    opsR = llq_1R.to_numpy().tolist()
    opsL = llq_1L.to_numpy().tolist()
    dudi = GenerateAssembly(p1, p2, p3, opsR)
    dodi = GenerateAssembly(p1, p2, p3, opsL)

    dudi.columns=['X1', 'Y1', 'Z1','X2','Y2','Z2']
    dudi.insert(0, 'GEO', '3DFACE')
    dudi.set_index('GEO',inplace= True)
    dudi["P1"] = dudi["X1"].map(str)+','+dudi["Y1"].map(str)+','+dudi["Z1"].map(str)
    dudi["P2"] = dudi["X2"].map(str)+','+dudi["Y2"].map(str)+','+dudi["Z2"].map(str)
    dudi['TERMIN'] = '{ENTER}'
    dudi.drop(['X1', 'Y1', 'Z1','X2','Y2','Z2'], inplace=True, axis=1)
    dudi.to_pickle(cathi+"/_$temp/"+ str(i)+".pkl") 
    
    dodi.columns=['X1', 'Y1', 'Z1','X2','Y2','Z2']
    dodi.insert(0, 'GEO', '3DFACE')
    dodi.insert(7,'TERMIN', '{ENTER}')
    dodi.set_index('GEO',inplace= True)
    dodi["P1"] = dodi["X1"].map(str)+','+dodi["Y1"].map(str)+','+dodi["Z1"].map(str)
    dodi["P2"] = dodi["X2"].map(str)+','+dodi["Y2"].map(str)+','+dodi["Z2"].map(str)
    dodi['TERMIN'] = '{ENTER}'
    dodi.drop(['X1', 'Y1', 'Z1','X2','Y2','Z2'], inplace=True, axis=1)
    dodi.to_pickle(cathi+"/_$temp/"+ str(i)+str(i)+".pkl")


for i in range(270,280,1):
    viva1 = pd.read_pickle(cathi+"/_$temp/"+str(i)+".pkl")
    viva2 = pd.read_pickle(cathi+"/_$temp/"+str(i+1)+".pkl")
    viva2 = viva2[['P2','P1']]
    viva3 = pd.concat([viva1, viva2], axis=1)
    
    wbdu = Workbook()
    wsdu = wbdu.active
    for r in dataframe_to_rows(viva3, index=True, header=True):
        wsdu.append(r)
    for cell in wsdu['A'] + wsdu[1]:
        cell.style = 'Pandas'
    
    wbbdu = openpyxl.load_workbook(generator1, keep_vba=True)
    sheet = wbbdu['Sheet1']
    for r in dataframe_to_rows(viva3, index=True, header=True):
        sheet.append(r)
    for cell in sheet['A'] + sheet[1]:
        cell.style = 'Pandas'
    wbbdu.save(generator1)
####   
for i in range(281,287,1):
    viva1 = pd.read_pickle(cathi+"/_$temp/"+str(i)+".pkl")
    viva2 = pd.read_pickle(cathi+"/_$temp/"+str(i+1)+".pkl")
    viva2 = viva2[['P2','P1']]
    viva3 = pd.concat([viva1, viva2], axis=1)
    
    wbdu = Workbook()
    wsdu = wbdu.active
    for r in dataframe_to_rows(viva3, index=True, header=True):
        wsdu.append(r)
    for cell in wsdu['A'] + wsdu[1]:
        cell.style = 'Pandas'
    
    wbbdu = openpyxl.load_workbook(generator1, keep_vba=True)
    sheet = wbbdu['Sheet1']
    for r in dataframe_to_rows(viva3, index=True, header=True):
        sheet.append(r)
    for cell in sheet['A'] + sheet[1]:
        cell.style = 'Pandas'
    wbbdu.save(generator1)
####   
for i in range(288,293,1):
    viva1 = pd.read_pickle(cathi+"/_$temp/"+str(i)+".pkl")
    viva2 = pd.read_pickle(cathi+"/_$temp/"+str(i+1)+".pkl")
    viva2 = viva2[['P2','P1']]
    viva3 = pd.concat([viva1, viva2], axis=1)
    
    wbdu = Workbook()
    wsdu = wbdu.active
    for r in dataframe_to_rows(viva3, index=True, header=True):
        wsdu.append(r)
    for cell in wsdu['A'] + wsdu[1]:
        cell.style = 'Pandas'
    
    wbbdu = openpyxl.load_workbook(generator1, keep_vba=True)
    sheet = wbbdu['Sheet1']
    for r in dataframe_to_rows(viva3, index=True, header=True):
        sheet.append(r)
    for cell in sheet['A'] + sheet[1]:
        cell.style = 'Pandas'
    wbbdu.save(generator1)
####   
for i in range(294,298,1):
    viva1 = pd.read_pickle(cathi+"/_$temp/"+str(i)+".pkl")
    viva2 = pd.read_pickle(cathi+"/_$temp/"+str(i+1)+".pkl")
    viva2 = viva2[['P2','P1']]
    viva3 = pd.concat([viva1, viva2], axis=1)
    
    wbdu = Workbook()
    wsdu = wbdu.active
    for r in dataframe_to_rows(viva3, index=True, header=True):
        wsdu.append(r)
    for cell in wsdu['A'] + wsdu[1]:
        cell.style = 'Pandas'
    
    wbbdu = openpyxl.load_workbook(generator1, keep_vba=True)
    sheet = wbbdu['Sheet1']
    for r in dataframe_to_rows(viva3, index=True, header=True):
        sheet.append(r)
    for cell in sheet['A'] + sheet[1]:
        cell.style = 'Pandas'
    wbbdu.save(generator1)
############################################################################
############################################################################
for i in range(270,280,1):
    viva1 = pd.read_pickle(cathi+"/_$temp/"+str(i)+str(i)+".pkl")
    viva2 = pd.read_pickle(cathi+"/_$temp/"+str(i+1)+str(i+1)+".pkl")
    viva2 = viva2[['P2','P1']]
    viva3 = pd.concat([viva1, viva2], axis=1)
    
    wbdu = Workbook()
    wsdu = wbdu.active
    for r in dataframe_to_rows(viva3, index=True, header=True):
        wsdu.append(r)
    for cell in wsdu['A'] + wsdu[1]:
        cell.style = 'Pandas'
    
    wbbdu = openpyxl.load_workbook(generator1, keep_vba=True)
    sheet = wbbdu['Sheet1']
    for r in dataframe_to_rows(viva3, index=True, header=True):
        sheet.append(r)
    for cell in sheet['A'] + sheet[1]:
        cell.style = 'Pandas'
    wbbdu.save(generator1)
####   
for i in range(281,287,1):
    viva1 = pd.read_pickle(cathi+"/_$temp/"+str(i)+str(i)+".pkl")
    viva2 = pd.read_pickle(cathi+"/_$temp/"+str(i+1)+str(i+1)+".pkl")
    viva2 = viva2[['P2','P1']]
    viva3 = pd.concat([viva1, viva2], axis=1)
    
    wbdu = Workbook()
    wsdu = wbdu.active
    for r in dataframe_to_rows(viva3, index=True, header=True):
        wsdu.append(r)
    for cell in wsdu['A'] + wsdu[1]:
        cell.style = 'Pandas'
    
    wbbdu = openpyxl.load_workbook(generator1, keep_vba=True)
    sheet = wbbdu['Sheet1']
    for r in dataframe_to_rows(viva3, index=True, header=True):
        sheet.append(r)
    for cell in sheet['A'] + sheet[1]:
        cell.style = 'Pandas'
    wbbdu.save(generator1)
####   
for i in range(288,293,1):
    viva1 = pd.read_pickle(cathi+"/_$temp/"+str(i)+str(i)+".pkl")
    viva2 = pd.read_pickle(cathi+"/_$temp/"+str(i+1)+str(i+1)+".pkl")
    viva2 = viva2[['P2','P1']]
    viva3 = pd.concat([viva1, viva2], axis=1)
    
    wbdu = Workbook()
    wsdu = wbdu.active
    for r in dataframe_to_rows(viva3, index=True, header=True):
        wsdu.append(r)
    for cell in wsdu['A'] + wsdu[1]:
        cell.style = 'Pandas'
    
    wbbdu = openpyxl.load_workbook(generator1, keep_vba=True)
    sheet = wbbdu['Sheet1']
    for r in dataframe_to_rows(viva3, index=True, header=True):
        sheet.append(r)
    for cell in sheet['A'] + sheet[1]:
        cell.style = 'Pandas'
    wbbdu.save(generator1)
####   
for i in range(294,298,1):
    viva1 = pd.read_pickle(cathi+"/_$temp/"+str(i)+str(i)+".pkl")
    viva2 = pd.read_pickle(cathi+"/_$temp/"+str(i+1)+str(i+1)+".pkl")
    viva2 = viva2[['P2','P1']]
    viva3 = pd.concat([viva1, viva2], axis=1)
    
    wbdu = Workbook()
    wsdu = wbdu.active
    for r in dataframe_to_rows(viva3, index=True, header=True):
        wsdu.append(r)
    for cell in wsdu['A'] + wsdu[1]:
        cell.style = 'Pandas'
    
    wbbdu = openpyxl.load_workbook(generator1, keep_vba=True)
    sheet = wbbdu['Sheet1']
    for r in dataframe_to_rows(viva3, index=True, header=True):
        sheet.append(r)
    for cell in sheet['A'] + sheet[1]:
        cell.style = 'Pandas'
    wbbdu.save(generator1)


   