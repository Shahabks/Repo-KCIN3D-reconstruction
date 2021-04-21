# Libraries 
import os
import pandas as pd
import numpy as np
import math
from pandas import DataFrame
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

# Section-1

# The users enter the datasets names (e.g. drawings numbers, projects) and a range of Tag numbers that represent the road-sections
## If we proceed to the production phase, we will store all drawings data from a project in one mega-database 
prnamepv = input ( "Enter the project plot name: " )
prnamelv = input ( "Enter the project profile name: " )
querytagA = int(input("Enter the 1st tag for DX: "))
querytagB = int(input("Enter the last tag for DX: "))

cathy = 'mnt/c/Users/CINPC0075/Desktop/Repo-KCIN3D-reconstruction/'
mdpathpv = cathy+ prnamepv +'.csv'
mdpathlv = cathy+ prnamelv +'.csv'
mdpathpv_OUT0 = cathy+ prnamepv+"_OUT0" +'.xlsx'
mdpathpv_OUT1 = cathy+ prnamepv+"_OUT1" +'.xlsx'


dflv = pd.read_csv(mdpathlv,index_col=0)
dfpv = pd.read_csv(mdpathpv,index_col=0)


# Section-2

# Here we do some arrangements, parameter-setting and cleaning the header
to_drop = ['Author','Closed', 'Comments', 'Drawing Revision Number', 'EdgeStyleId', 'FaceStyleId', 'File Accessed', 'File Created', 'File Last Saved By', 'File Location',
           'File Modified','File Name', 'File Size', 'Fit/Smooth', 'Hyperlink', 'Hyperlink Base', 'Keywords']
dfpv.drop(to_drop, inplace=True, axis=1)
dflv.drop(to_drop, inplace=True, axis=1)

dfpv.set_index("Value",inplace= True)
dflv.set_index("Value",inplace= True)
rangi = querytagB - querytagA

# Section-3

# Extracting the coordinates x,y associated with Tag-Numbers (e.g. NO.270). The range has been specified by the users in Section-1, line-3 and 4
dtag = []
ran = range(querytagA,querytagB)
for i in ran:
    m = "NO."+str(i)
    dtag.append(m)
row = dtag

dff = dfpv.loc[row, ["Position X", "Position Y", "Position Z"]]
dff.sort_values("Position X", inplace=True, ascending=True)
## dff contains X,Y of each Tag-number ............... For details, see Work_ON_DX_Komatsu.ipynb
### the information in dff will help us find the coordinates of each point along the road-centerline 
#### Tag-numbers are major keys for us to connect different drawings together and extract relevant information


# Section-4

# Here, we do Euclidean distance similarity measure to find the pair (Tag-number, corresponding-road-centerline-point)
## We also extract the pair (Tag-Number, corresponding-circle-along-road-centerline)
df1 = dfpv.loc[dfpv["Name"] == 'Circle']
df2 = df1.loc[df1["Radius"] == 2.5]
df22 = df2.loc[df2["Name"] == "Circle", ["Center X", "Center Y", "Center Z"]]
df22.sort_values("Center X", inplace=True, ascending=True)
rrx = df22["Center X"]
rry = df22["Center Y"]
rrrx = dff["Position X"]
rrry = dff["Position Y"]
rrrr0 = dff.index
import math
fel=[]
for i, j in zip(rrx,rry):
    for m, n, k in zip(rrrx,rrry,rrrr0):
        f= math.sqrt((j-n)**2 + (i-m)**2)
        fel.append([i,j,m,n,k,f])

from pandas import DataFrame
d1 = pd.DataFrame(fel, columns=['CX','CY','TX','TY','TG','Dis'])
d11 = d1[:30]
d110 = d11.sort_values(by='Dis', ascending=True)
d11out=d110[:1]
d12 = d1[30:60]
d120 = d12.sort_values(by='Dis', ascending=True)
d12out= d120[:1]
d13 = d1[60:90]
d130 = d13.sort_values(by='Dis', ascending=True)
d13out = d130[:1]
d14 = d1[90:120]
d140 = d14.sort_values(by='Dis', ascending=True)
d14out = d140[:1]
d15 = d1[120:150]
d150 = d15.sort_values(by='Dis', ascending=True)
d15out = d150[:1]
d16 = d1[150:180]
d160 = d16.sort_values(by='Dis', ascending=True)
d16out = d160[:1]
d17 = d1[180:210]
d170 = d17.sort_values(by='Dis', ascending=True)
d17out = d170[:1]
dframe =  [d11out,d12out,d13out,d14out,d15out,d16out,d17out]
fi = pd.concat(dframe)
fi.reset_index(drop=True, inplace=True)
fi.sort_values("CX", inplace=True, ascending=True)
fi.insert(0, 'GEO', 'Circle')
fi.set_index('GEO',inplace= True)
fi["P1"] = fi["CX"].map(str)+','+fi["CY"].map(str)
fi.insert(7, 'P2', '1.5')
dl1 = dfpv.loc[dfpv["Name"] == 'Line']
dl2 = dl1.loc[dl1['Color'] == "red"]
dl3 = dl2.loc[dl2["Length"] == 5]
dl4 = dl3.loc[dl3['Name'] == "Line", ["End X", "End Y", "End Z",'Start X','Start Y','Start Z']]
dl4.reset_index(drop=True, inplace=True)
dl4.sort_values("End X", inplace=True, ascending=True)
dl4.insert(0, 'GEO', 'Line')
dl4.set_index('GEO',inplace= True)
dl4["P1"] = dl4["End X"].map(str)+','+dl4["End Y"].map(str)+','+dl4["End Z"].map(str)
dl4["P2"] = dl4["Start X"].map(str)+','+dl4["Start Y"].map(str)+','+dl4["Start Z"].map(str)
dl4l=dl9l = dl4[:rangi]
dl4l.insert(8, 'Value', row)
dll4 = dl4.drop(["End X", "End Y", "End Z",'Start X','Start Y','Start Z'], axis=1)
fii = fi.drop(["CX", "CY", "TX",'TY','TG','Dis'], axis=1)
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(dll4, index=True, header=True):
    ws.append(r)
for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'
wb.save(mdpathpv_OUT0)
wba = Workbook()
wsa = wba.active
for k in dataframe_to_rows(fii, index=True, header=True):
    wsa.append(k)
for cell in wsa['A'] + wsa[1]:
    cell.style = 'Pandas'
wba.save(mdpathpv_OUT1)

##### The ultimate outputs of Sectio-4 are the coordinations (X,Y,Z) of all designated points along the road-centerline (Stored in a xlsx file for re-creating and 
#####   drawing the road-centerline by use of VBA)
##### RECALL, The true value of Z for the points must be extracted from Profile-view, so we will find the true elevation and replace Z with the true value 
##### RECALL, we use Tag-numbers to extract the true elevation and value of Z for each datapoint through the following sections 


# Section-5

# here, we use the Tag-numbers information and Profile-view drawings to extract the true values for Z

## these lines of code help us find the box in the drawings that contains true elevations
ref0 = dflv.loc[["高", "計","画"], ["Position X", "Position Y", "Position Z"]]
r1 = ref0['Position Y'].min()
r2 = ref0['Position Y'].max()
lm = dflv.loc[dflv["Position Y"]>(r1+8.491)]
lq = lm.loc[lm["Position Y"]<(r1+22.491)]
lz = lq.loc[:,["Position X","Position Y","Position Z"]]
dtemp1 = dflv.reset_index()
dtemp2 = dtemp1.dropna(subset=['Value'])

## we get the grand reference information
dtemp3 = dtemp2[dtemp2['Value'].str.contains('DL=')]
dtemp4 = dtemp3.iat[0,0]
reflev = float(dtemp4.replace('DL=',''))


## we extract the true elevations for each point
lz1 = lz.drop(["高", "計","画"])
lz2 = lz1.reset_index()
lz2.sort_values('Position X', inplace=True, ascending=True)
lz22 = lz2.loc[:,["Position X"]]
lz2['Value'] = lz2['Value'].astype(float)
lz3 = lz2['Value'] - reflev 
lz22.insert(0, 'TRUE ELEVATION', lz3)
b1 = lz22['Position X'].min()
b2 = lz22['Position X'].max()

## we extract and concatenate the true elevations with the corresponding Tag-numbers
dtemp33 = dtemp2[dtemp2['Value'].str.contains('NO.')]
dtemp333 = dtemp33.loc[dtemp33['Position X']>= b1]
dtemp444 = dtemp333.loc[dtemp333['Position X']<= b2]
dtemp444.sort_values('Position X', inplace=True, ascending=True)
dTG0 = dtemp444.loc[:,['Value','Position X']]
dTG = dTG0.loc[dTG0["Value"].apply(len)<7]
z1 = lz22.set_index("Position X")
z2 = dTG.set_index("Position X")
dTGLEV=pd.merge(z1, z2, left_index=True, right_index=True, how='outer')
tt = dTGLEV.dropna()
dimi = tt.loc[tt['Value']>=str(querytagA)]
dimit = dimi.loc[dimi['Value']<str(querytagB)]
dimit=dimit[['Value','TRUE ELEVATION']]
dimit
dl9l = dl4l[:23]


print(dl9l,tt)